"""
01_load_raw.py — Parse all 3 XLSX files into clean wide-format CSVs.

Each XLSX stores data in a long/grouped format where each frequency record
spans multiple rows (one row per dimension). This script flattens them into
one row per design point.

Usage:
    python code/data/01_load_raw.py
"""

import re
import logging
import argparse
import numpy as np
import pandas as pd
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger(__name__)

# Allow running from any working directory
import sys
_code_dir = next(p for p in Path(__file__).resolve().parents if p.name == "code")
sys.path.insert(0, str(_code_dir))

from utils.config import DATA_PROC_DIR, ROOT


# ─────────────────────────────────────────────────────────────────────────────
# Helper: extract numeric value from strings like "42.8 mm", "1 GHz"
# ─────────────────────────────────────────────────────────────────────────────

def _parse_num(val) -> float | None:
    """Extract first numeric value from a cell (string or number)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    m = re.search(r"[-+]?\d*\.?\d+", str(val))
    return float(m.group()) if m else None


def _parse_freq_ghz(val) -> float | None:
    """Parse frequency cell → GHz. Handles '1 GHz', '2.454', '12 GHz'."""
    if val is None:
        return None
    txt = str(val).strip()
    num = _parse_num(txt)
    if num is None:
        return None
    # Already in GHz (values are given as e.g. '1 GHz', '12 GHz')
    return float(num)


# ─────────────────────────────────────────────────────────────────────────────
# Dimension name → column name mapping
# ─────────────────────────────────────────────────────────────────────────────

_DIM_MAP = {
    "r1": ["outer radius of outer", "r1"],
    "r2": ["inner radius of outer", "r2"],
    "r3": ["outer radius of inner", "r3"],
    "r4": ["inner radius of inner", "r4"],
    "t":  ["thickness", "(t)"],
    "d1": ["d1"],
    "d2": ["d2"],
    "h":  ["height", " h "],
    "p":  ["the separation between", "period", "(p)"],
}


def _match_dim(text: str) -> str | None:
    """Return the canonical column name for a dimension label cell."""
    if not text:
        return None
    low = str(text).lower().strip()
    for col, patterns in _DIM_MAP.items():
        for pat in patterns:
            if pat.lower() in low:
                return col
    return None


# ─────────────────────────────────────────────────────────────────────────────
# Parser for unit-cell XLSX (0° and 180°)
# ─────────────────────────────────────────────────────────────────────────────

def parse_unit_cell_xlsx(fpath: Path, rotation: int) -> pd.DataFrame:
    """
    Parse 'Updated AI_ML Data.xlsx' or 'Updated AI_ML Data (1).xlsx'.

    Each design point spans 5+ consecutive rows:
      Row with Sl.No + Frequency → start of new record
      Subsequent rows → dimension name + value pairs

    Parameters
    ----------
    fpath    : Path to the XLSX file
    rotation : 0 or 180 (used to label the rotation column)

    Returns
    -------
    DataFrame with columns: r1,r2,r3,r4,t,d1,d2,h,p,freq_ghz,rotation
    """
    import openpyxl
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active

    records = []
    current: dict = {}
    current_freq: float | None = None

    for row in ws.iter_rows(values_only=True):
        # Strip all None-only rows
        vals = [v for v in row if v is not None]
        if not vals:
            continue

        row_str = " | ".join(str(v) for v in row)

        # ── Detect start of a new record (row contains a serial number + freq)
        # Look for a numeric Sl.No (int-like) and a GHz frequency in same row
        sl_found, freq_found = None, None
        for v in row:
            if isinstance(v, (int, float)) and 1 <= v <= 300 and sl_found is None:
                sl_found = v
            if v is not None and "ghz" in str(v).lower():
                freq_found = _parse_freq_ghz(v)
            if isinstance(v, str) and re.match(r"^\d+\s*GHz$", v.strip(), re.I):
                freq_found = _parse_freq_ghz(v)

        # Also handle case where freq is purely numeric and sl is in same row
        if sl_found is not None:
            for v in row:
                if isinstance(v, (int, float)) and 0.5 <= v <= 20 and v != sl_found:
                    freq_found = freq_found or float(v)

        if sl_found is not None and freq_found is not None:
            # Save previous record
            if current and current_freq is not None:
                current["freq_ghz"] = current_freq
                current["rotation"] = rotation
                records.append(dict(current))
            # Start new record
            current = {}
            current_freq = freq_found
            continue

        # ── Parse dimension rows
        # Scan every pair of (dim_label_cell, value_cell)
        row_list = list(row)
        for i in range(len(row_list) - 1):
            col = _match_dim(row_list[i])
            if col and col not in current:
                val = _parse_num(row_list[i + 1])
                if val is not None:
                    current[col] = val

    # Save last record
    if current and current_freq is not None:
        current["freq_ghz"] = current_freq
        current["rotation"] = rotation
        records.append(dict(current))

    df = pd.DataFrame(records)
    log.info(f"Parsed {len(df)} records from {fpath.name} (rotation={rotation}°)")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Parser for full-structure XLSX
# ─────────────────────────────────────────────────────────────────────────────

_FULL_DIM_MAP = {
    "L":     ["waveguide couplers length (l)", "coupler length"],
    "r_wg":  ["inner radius of waveguide coupler", "radius of waveguide"],
    "P_wg":  ["length of the circular waveguide"],
    "R_uc":  ["inner radius of cesrr unit cell"],
    "p_sep": ["separation between two cesrr", "(p)"],
    "D":     [" d "],
}

_UC_DIM_MAP_FULL = {
    "r1": ["outer radius of outer"],
    "r2": ["inner radius of outer"],
    "r3": ["outer radius of inner"],
    "r4": ["inner radius of inner"],
    "t":  ["thickness"],
}


def parse_full_structure_xlsx(fpath: Path) -> pd.DataFrame:
    """
    Parse 'Full_structure_dimensions.xlsx' which has two side-by-side tables:
    Left table: full-structure waveguide dimensions
    Right table: unit-cell dimensions
    """
    import openpyxl
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active

    records = []
    current_full: dict = {}
    current_uc: dict = {}
    current_freq_full: float | None = None
    current_freq_uc: float | None = None
    prev_sl = None

    for row in ws.iter_rows(values_only=True):
        row_list = list(row)
        if all(v is None for v in row_list):
            continue

        # Detect new record by serial number in either left or right table
        # Left table occupies cols 1-5, right table cols 6-10 (approx)
        left  = row_list[:5]
        right = row_list[5:]

        # Look for freq in left half
        for v in left:
            if v is not None and str(v).strip().lower().endswith("ghz"):
                current_freq_full = _parse_freq_ghz(v)
            if isinstance(v, (int, float)) and 0.5 <= float(v) <= 30:
                pass  # could be freq numeric

        # Look for freq in right half
        for v in right:
            if v is not None and str(v).strip().lower().endswith("ghz"):
                current_freq_uc = _parse_freq_ghz(v)

        # Detect Sl.No + freq → new record
        sl_vals = [v for v in row_list if isinstance(v, (int, float)) and 1 <= v <= 100]
        freq_vals = [_parse_freq_ghz(v) for v in row_list
                     if v is not None and str(v).strip().lower().endswith("ghz")]

        if sl_vals and freq_vals:
            if prev_sl is not None and prev_sl != sl_vals[0]:
                # Save previous
                merged = {**current_full, **current_uc}
                if current_freq_full is not None:
                    merged["freq_ghz"] = current_freq_full
                    records.append(merged)
                current_full = {}
                current_uc = {}
            prev_sl = sl_vals[0] if sl_vals else prev_sl

        # Parse left-side dims (full structure)
        for i in range(len(left) - 1):
            low = str(left[i]).lower() if left[i] else ""
            for col, pats in _FULL_DIM_MAP.items():
                if any(p in low for p in pats) and col not in current_full:
                    val = _parse_num(left[i + 1])
                    if val is not None:
                        current_full[col] = val

        # Parse right-side dims (unit cell within full structure)
        for i in range(len(right) - 1):
            low = str(right[i]).lower() if right[i] else ""
            for col, pats in _UC_DIM_MAP_FULL.items():
                if any(p in low for p in pats) and col not in current_uc:
                    val = _parse_num(right[i + 1])
                    if val is not None:
                        current_uc[col] = val

    # Last record
    if current_full or current_uc:
        merged = {**current_full, **current_uc}
        if current_freq_full is not None:
            merged["freq_ghz"] = current_freq_full
            records.append(merged)

    df = pd.DataFrame(records)
    log.info(f"Parsed {len(df)} full-structure records from {fpath.name}")
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # Locate raw files (look in DATA_RAW_DIR first, then repo root)
    def _find(fname):
        for d in [ROOT / "data" / "raw", ROOT]:
            p = d / fname
            if p.exists():
                return p
        raise FileNotFoundError(f"Cannot find raw file: {fname}")

    DATA_PROC_DIR.mkdir(parents=True, exist_ok=True)

    # --- Unit cell 0° ---
    df_0 = parse_unit_cell_xlsx(_find("Updated AI_ML Data.xlsx"), rotation=0)
    out_0 = DATA_PROC_DIR / "raw_unit_0deg.csv"
    df_0.to_csv(out_0, index=False)
    log.info(f"Saved: {out_0} ({len(df_0)} rows)")

    # --- Unit cell 180° ---
    df_180 = parse_unit_cell_xlsx(_find("Updated AI_ML Data (1).xlsx"), rotation=180)
    out_180 = DATA_PROC_DIR / "raw_unit_180deg.csv"
    df_180.to_csv(out_180, index=False)
    log.info(f"Saved: {out_180} ({len(df_180)} rows)")

    # --- Full structure ---
    df_full = parse_full_structure_xlsx(_find("Full_structure_dimensions.xlsx"))
    out_full = DATA_PROC_DIR / "raw_full_structure.csv"
    df_full.to_csv(out_full, index=False)
    log.info(f"Saved: {out_full} ({len(df_full)} rows)")

    # --- Combined unit cell (both rotations) ---
    df_combined = pd.concat([df_0, df_180], ignore_index=True)
    out_comb = DATA_PROC_DIR / "raw_unit_combined.csv"
    df_combined.to_csv(out_comb, index=False)
    log.info(f"Saved combined: {out_comb} ({len(df_combined)} rows)")

    print("\n✅ Step 01 complete. Files written to:", DATA_PROC_DIR)


if __name__ == "__main__":
    main()
